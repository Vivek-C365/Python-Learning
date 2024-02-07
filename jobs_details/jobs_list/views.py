from django.shortcuts import render, redirect , HttpResponse,get_object_or_404
from django.contrib.auth import  logout
from django.contrib.auth.decorators import login_required
from .models import adm , indeed_job_list ,Source , Company
from openpyxl import load_workbook
# Create your views here.
def ad(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        print(username, email,password)
        us = adm(username = username, email=email,password=password)
        us.save()
        
        return HttpResponse(' Admin is stored')  #redirecting to the home page after registration
        
    return render(request,'ad.html')
def login_view(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get("passwd")

        # Use Django's ORM to query the database
        result = adm.objects.filter(username=username, password=password)
        print(result)
        if result:
            # Authentication successful
            request.session['username'] = username
            return redirect('/dashboard/')
        else:
            # Authentication failed
            return redirect('/404/')
    return render(request, 'login.html')

def dashboard(request):
    context = {}
    # Check if the user is authenticated
    username = request.session.get('username')
     # Query all jobs from the database
    jobs = indeed_job_list.objects.all()
    # Get the count of entries in the Job model
    entry_count = indeed_job_list.objects.count()

    if not username:
        return redirect('/login/')   #Redirect to Login Page if User Not logged in
    else:
        context["username"] = username     #Adding Logged In User Name To The Context Dictionary
        context["job_count"] = entry_count
        return render(request, 'dashboard.html',context)

def job_list(request):
    jobs = indeed_job_list.objects.all()
    
    if request.method == "POST":
        if 'delete' in request.POST:
            job_id_to_delete = request.POST.get('delete')
            job_to_delete = indeed_job_list.objects.get(pk=job_id_to_delete)
            job_to_delete.delete()
            # return redirect('job_list')# Redirect to the same page after deletion
        else:
            # Update operation
            job_id_to_update = request.POST.get('update')
            job_to_update = indeed_job_list.objects.get(pk=job_id_to_update)
            print(job_id_to_update,job_to_update)
            job_to_update.job_title = request.POST.get('job_title')
            job_to_update.company = request.POST.get('company')
            job_to_update.location = request.POST.get('location')
            job_to_update.salary_range = request.POST.get('salary_range')
            job_to_update.tags = request.POST.get('tags')
            job_to_update.save()
        # else:
        #     excel_file = request.FILES['excel_file']
        #     file_name = excel_file.name
        #     file = file_name.split('_')[0]
        #     if file_name.endswith('.xlsx'):
        #         if Source.objects.filter(name=file).exists():
        #             print('File already exists')
        #         else:
        #             Source.objects.create(name=file)
        #             print('File name Saved')
        #         wb = load_workbook(excel_file)
        #         wb = wb.active
        #         for row in wb.iter_rows(values_only=True):
        #             job_title, company_name, location, salary_range, tags = row
        #             indeed_job_list.objects.create(
        #                 job_title=job_title, 
        #                 company=company_name,
        #                 location=location,
        #                 salary_range=salary_range,
        #                 tags=tags
        #             )
        #         print(file)
        
        # return redirect('job_list')  # Redirect to the same page after adding jobs
    
    return render(request, 'job_list.html', {'jobs': jobs})


def LogoutPage(request):
    logout(request)
    return redirect('/')

def error_view(request):
    return render(request, '404.html')

def test(request):
    return render(request, 'test.html')

def indeed_job(request):
    if request.method == 'POST':
        custom_id = request.POST.get('custom_id')
        job_title = request.POST.get('job_title')
        company = request.POST.get('company')
        location = request.POST.get('location')
        salary_range = request.POST.get('salary_range')
        tags = request.POST.get('tags')
        date_of_post = request.POST.get('date_of_post')
        job_description = request.POST.get('job_description')
        # print(custom_id, job_title,company)
        us = indeed_job_list(custom_id = custom_id, job_title=job_title,company=company,location=location,salary_range=salary_range,tags=tags,date_of_post=date_of_post,job_description=job_description)
        us.save()
        
        return redirect('/dashboard/') 
        
    return render(request, 'indeed_job.html')

def Source_view(request):
    if request.method == 'POST':
        name= request.POST.get('name')
        source_fields =Source(name = name)
        source_fields.save()
    return render(request,  "source.html")

def Company_view(request):
    if request.method == 'POST':
        company_name = request.POST.get('Company_name')
        company_website = request.POST.get('Company_website')
        company_address = request.POST.get('Company_address')
        company_contact = request.POST.get('Company_contact')
        company_email = request.POST.get('Company_email')
        company_size = request.POST.get('Company_size')
        source_id = request.POST.get('source_id')
        company = Company(company_name=company_name,company_website=company_website,company_address=company_address,company_contact=company_contact,company_email=company_email,company_size=company_size,source_id=source_id)
        company.save()
    return redirect("/dashboard/")

